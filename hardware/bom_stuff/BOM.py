from ast import Try
import time
import logging

from urllib import request, parse
from pathlib import Path

import openpyxl as opxl
import bs4

import json

import gzip
import io

logger = logging.getLogger('BOM')
logger.setLevel(logging.DEBUG)
formatter = logging.Formatter('%(asctime)s - %(name)s - %(funcName)s - %(levelname)s - %(message)s')
ch = logging.StreamHandler()
ch.setLevel(logging.INFO)
ch.setFormatter(formatter)
logger.addHandler(ch)
fh = logging.FileHandler(Path(__file__ + '.log'))
fh.setLevel(logging.DEBUG)
fh.setFormatter(formatter)
logger.addHandler(fh)

FILE = 'BOM for ChipShouter-PicoEMP-NL.xlsx'
SHEET = 'Bill of Materials-ChipS SMD'

FETCH_ONLINE = False

col_name = "B"
col_description = "C"
col_quantity = "E"
col_partnumber = "G"
col_alt = "U"

col_digikey = 'K'
col_mouser = 'P'

wb = opxl.load_workbook(Path(__file__).parent / FILE)
ws = wb[SHEET]

headers = {
    'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:96.0) Gecko/20100101 Firefox/96.0',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
    'Accept-Encoding': 'gzip, deflate, br',
}

def parse_digikey(resp):
    assert(resp.info()["Content-Encoding"] == 'gzip')

    buf = io.BytesIO(resp.read())
    data = gzip.GzipFile(fileobj=buf)
    text = data.read()
    soup = bs4.BeautifulSoup(text, "html.parser")
    info_text = soup.find(id="__NEXT_DATA__")
    assert(info_text.get('type') == 'application/json')

    stock = None
    price_1 = None
    price_10 = None

    info_dict = json.loads(info_text.string)
    stock = info_dict["props"]["pageProps"]["envelope"]["data"]["priceQuantity"].get("qtyAvailable")
    logger.debug(f'digikey stock raw: {stock}')
    if stock:
        stock = int(stock.replace('.', '_'))

    pricings = info_dict["props"]["pageProps"]["envelope"]["data"]["priceQuantity"].get("pricing")
    logger.debug(f'digikey pricings raw: {pricings}')
    if pricings:
        for pricing in pricings:
            packaging = pricing.get("packaging")
            if packaging and ('CT' in packaging or 'Tray' in packaging or 'Bulk' in packaging):
                quantities = pricing.get("pricingTiers")
                for quantity in quantities:
                    if quantity['breakQty'] == '1':
                        price_1 = float(quantity['extendedPrice'][2:].replace(',', '.'))
                    if quantity['breakQty'] == '10':
                        price_10 = float(quantity['extendedPrice'][2:].replace(',', '.'))
                # print(pricing.get("pricingTiers"))
    
    ret = {'stock': stock, 'price_1': price_1, 'price_10': price_10}
    logger.info(f'digikey: {ret}')
    return ret

def parse_mouser(resp):
    assert(resp.info()["Content-Encoding"] == 'gzip')

    buf = io.BytesIO(resp.read())
    data = gzip.GzipFile(fileobj=buf)
    text = data.read()
    soup = bs4.BeautifulSoup(text, "html.parser")

    stock = None
    price_1 = None
    price_10 = None

    # var productUnitPricing
    stock_tag = soup.find("h2", class_='pdp-pricing-header')
    logger.debug(f'mouser stock_tag: {stock_tag}')
    assert(stock_tag != None)
    if 'voorraad' in stock_tag.string:
        stock = stock_tag.string.split("voorraad: ")[1].split("\r")[0]
        stock = int(stock.replace('.', '_'))
    # print(f'stock: {stock}')
    
    # Add a space and pray that it is never the last one
    price_1_tag = soup.find("td", headers=lambda x: x and "pricebreakqty_1 " in x and "unitpricecolhdr" in x)
    logger.debug(f'mouser price_1_tag: {price_1_tag}')
    if price_1_tag:
        price_1 = price_1_tag.string.split("€ ")[1].split("\r")[0]
        price_1 = float(price_1.replace(',', '.'))
    
    price_10_tag = soup.find("td", headers=lambda x: x and "pricebreakqty_10 " in x and "unitpricecolhdr" in x)
    logger.debug(f'mouser price_10_tag: {price_10_tag}')
    if price_10_tag:
        price_10 = price_10_tag.string.split("€ ")[1].split("\r")[0]
        price_10 = float(price_10.replace(',', '.')) * 10

    ret = {'stock': stock, 'price_1': price_1, 'price_10': price_10}
    logger.info(f'mouser: {ret}')
    return ret

def get_part_info(url, parse_f):
    req = request.Request(url, headers=headers)
    logger.debug(req.full_url)
    resp = request.urlopen(req, timeout=5)
    logger.debug(f'resp.status: {resp.status}')
    assert(resp.status == 200)
    logger.debug(f'resp encoding: {resp.info()["Content-Encoding"]}')
    return parse_f(resp)

if FETCH_ONLINE:
    parts_info = {}

    for row in range(2,1000):
        name = ws[f'{col_name}{row}'].value
        if not name:
            break
        description = ws[f'{col_description}{row}'].value
        partnumber = ws[f'{col_partnumber}{row}'].value
        quantity = ws[f'{col_quantity}{row}'].value
        url_digikey = ws[f'{col_digikey}{row}'].value
        url_mouser = ws[f'{col_mouser}{row}'].value
        logger.info(f'row: {row}, name: {name}, partnumber: {partnumber}, quantity: {quantity}')
        logger.info(f'description: {description}')

        logger.info(f'url digikey: {url_digikey}')
        data_digikey = None
        if url_digikey:
            # digikey doesnt always cooperate
            for _ in range(5):
                try:
                    data_digikey = get_part_info(url_digikey, parse_digikey)
                    data_digikey['url'] = url_digikey
                    break
                except Exception as e:
                    logger.exception(e)
                    logger.warning('Trying again')
        
        logger.info(f'url mouser : {url_mouser}')
        data_mouser = None
        if url_mouser:
            try:
                data_mouser = get_part_info(url_mouser, parse_mouser)
                data_mouser['url'] = url_mouser
            except Exception as e:
                logger.exception(e)

        parts_info[partnumber] = {
            'name': name,
            'quantity': quantity,
            'xlsxrow': row,
            'digikey': data_digikey,
            'mouser' : data_mouser
        }

        print('throttle to 1 per 1 second')
        time.sleep(1)

    with open(Path(__file__).parent / 'parts_info.json', 'w') as f:
        print(json.dumps(parts_info, indent=2), file=f)

else:
    parts_info = {}
    with open(Path(__file__).parent / 'parts_info.json', 'r') as f:
        parts_info = json.loads(f.read())

list_mouser = []
list_digikey = []
list_alt = []

total_mouser = 0
total_digikey = 0

for partnumber, info in parts_info.items():
    logger.info(partnumber)
    price_mouser = None
    if info.get('mouser') and info.get('mouser').get('stock'):
        logger.info(f"mouser stock: {info.get('mouser').get('stock')}")
        price_mouser = info.get('mouser').get('price_10')
    price_digikey = None
    if info.get('digikey') and info.get('digikey').get('stock'):
        logger.info(f"digikey stock: {info.get('digikey').get('stock')}")
        price_digikey = info.get('digikey').get('price_10')
    
    if price_digikey and price_mouser:
        if price_mouser < price_digikey:
            logger.info(f"mouser cheaper ({price_mouser} < {price_digikey})")
            logger.debug(info.get('mouser').get('url'))
            list_mouser.append(f"{info.get('mouser').get('url')} ; Quantity: {info['quantity'] * 10} ; Price: {price_mouser * info['quantity']}")
            total_mouser += price_mouser * info['quantity']
        else:
            logger.info(f"digikey cheaper ({price_digikey} < {price_mouser})")
            logger.debug(info.get('digikey').get('url'))
            list_digikey.append(f"{info.get('digikey').get('url')} ; Quantity: {info['quantity'] * 10} ; Price: {price_digikey * info['quantity']}")
            total_digikey += price_digikey * info['quantity']
    elif not price_digikey and price_mouser:
        logger.info(f"mouser stock only ({price_mouser})")
        logger.debug(info.get('mouser').get('url'))
        list_mouser.append(f"{info.get('mouser').get('url')} ; Quantity: {info['quantity'] * 10} ; Price: {price_mouser * info['quantity']}")
        total_mouser += price_mouser * info['quantity']
    elif price_digikey and not price_mouser:
        logger.info(f"digikey stock only ({price_digikey})")
        logger.debug(info.get('digikey').get('url'))
        list_digikey.append(f"{info.get('digikey').get('url')} ; Quantity: {info['quantity'] * 10} ; Price: {price_digikey * info['quantity']}")
        total_digikey += price_digikey * info['quantity']
    else:
        logger.info(f"need alt for {partnumber} ({info['name']})")
        alt = ws[f"{col_alt}{info['xlsxrow']}"].value
        if alt:
            logger.info(f'suggestion: {alt}')
        list_alt.append(f"{partnumber} ; Quantity: {info['quantity'] * 10} ; suggestion: {alt}")

with open(Path(__file__).parent / 'parts_shoppinglist.txt', 'w') as f:
    print("DIGIKEY", file=f)
    for url in list_digikey:
        print(url, file=f)
    print(f'total: {total_digikey}', file=f)

    print('', file=f)

    print("MOUSER", file=f)
    for url in list_mouser:
        print(url, file=f)
    print(f'total: {total_mouser}', file=f)

    print('', file=f)
    print("ALT", file=f)
    for part in list_alt:
        print(part, file=f)
